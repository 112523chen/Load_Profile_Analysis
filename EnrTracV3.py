from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook, Workbook
import pandas as pd
import os
import sys


def get_college_info(path):
    workbook = path
    wb = load_workbook(workbook)
    ws = wb['Data']
    info = ws["A1"].value
    school_name = info.split(" -")[0]
    return school_name


def get_account_info(path):
    workbook = path
    wb = load_workbook(workbook)
    ws = wb['Data']
    row = list(ws["2"])
    row = row[4:]
    row = [c.value for c in row]
    row = [r.split(" - ")[1] for r in row]
    if(len(set(row)) != 1):
        raise Exception(
            f"In file {path.split('/')[1]} multiple accounts exists. \nPlease make sure the file in {path.split('/')[1]} only contains one account")
    return list(set(row))[0]


def get_data(path):
    #Creates temp excel workbook
    workbook = path

    #pulls data from excel file
    wb = load_workbook(workbook)
    ws = wb['Data']

    #removes first row of worksheet
    points = list()
    row_count = ws.max_row
    ws.delete_rows(0, 1)
    ws.delete_rows((row_count-1), 1)

    #creates dataframe from the rows of the worksheet
    for i in ws.values:
        points.append(i)
    df = pd.DataFrame(points)
    df.columns = df.iloc[0]

    #deletes machine energy consumptions columns and creates a new column for the collective energy consumption
    c = list(df.columns)[0:4]
    for i in range(4, len(df.columns)):
        c.append(str(i))
    df.columns = c
    df["total"] = df[c[4:]].sum(axis=1)
    c.append("total")
    targets = c[:4]
    targets.append("total")
    df = df[(targets)]
    df = df.iloc[1:]

    #adds columns for references within code about date and time
    df["year"] = pd.DatetimeIndex(df["Interval End"]).year
    df["month"] = pd.DatetimeIndex(df["Interval End"]).month
    df["day"] = pd.DatetimeIndex(df["Interval End"]).day
    df["hour"] = pd.DatetimeIndex(df["Interval End"]).hour

    #creates points for reference to find row for report: tuple(year,month,energy_consumption)
    values = (df.groupby(["year", "month"])["total"].max().tolist())
    points = list(df.groupby(["year", "month"]).groups.keys())
    for count, point in enumerate(points):
        points_ = list(point)
        points_.append(values[count])
        points[count] = tuple(points_)

    #Gathers date for row in file
    school_name = get_college_info(path)
    account_info = get_account_info(path)
    data_points = list()
    for point in points:
        data = df[(df.year == point[0]) & (
            df.month == point[1]) & (df.total == point[2])]

        school = school_name
        account = account_info
        date = f"{data.month.values[0]}-{data.day.values[0]}-{data.year.values[0]}"
        peak = data.total.values[0]

        data_points.append([school,account,date,peak])
    return data_points


#Column Names
headers = ['college', 'account', 'date', 'peak(Kw)']
#File Name
workbook_name = 'data.xlsx'
#Create Excel file
wb = Workbook()
page = wb.active
#Sheet Name
page.title = 'data'
#Add column names to Excel file
page.append(headers)

#finds file in a director and gets data
d = dict()
# root = "inputs"
root = list((sys.argv))[1]

for f in os.listdir(root):
    
    if(f == ".DS_Store"):
        continue

    path = f"{root}/{f}"
    data_points = get_data(path)

    for point in data_points:
        if(point[0] not in list(d.keys())):
            d[point[0]] = [point]
        else:
            (d[point[0]]).append(point)

for college in list(d.keys()):
    for point in d[college]:
        page.append(point)

wb.save(filename=workbook_name)